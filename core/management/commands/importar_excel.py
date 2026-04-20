"""Importa los datos reales del Libro de Bodega desde el Excel."""
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand


def abrir_libro(path):
    """Abre .xls con xlrd o .xlsx/.ods con openpyxl y devuelve un wrapper uniforme."""
    if path.endswith(".xls"):
        import xlrd
        return XlrdWorkbook(xlrd.open_workbook(path))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        return OpenpyxlWorkbook(wb)


class XlrdWorkbook:
    def __init__(self, wb): self._wb = wb
    def sheet_names(self): return self._wb.sheet_names()
    def sheet_by_name(self, name): return XlrdSheet(self._wb.sheet_by_name(name))


class XlrdSheet:
    def __init__(self, sh): self._sh = sh
    @property
    def nrows(self): return self._sh.nrows
    @property
    def ncols(self): return self._sh.ncols
    def cell_value(self, r, c): return self._sh.cell_value(r, c)


class OpenpyxlWorkbook:
    def __init__(self, wb): self._wb = wb
    def sheet_names(self): return self._wb.sheetnames
    def sheet_by_name(self, name): return OpenpyxlSheet(self._wb[name])


class OpenpyxlSheet:
    def __init__(self, ws):
        self._rows = list(ws.iter_rows(values_only=True))
    @property
    def nrows(self): return len(self._rows)
    @property
    def ncols(self): return max((len(r) for r in self._rows), default=0)
    def cell_value(self, r, c):
        try:
            v = self._rows[r][c]
            return v if v is not None else ""
        except IndexError:
            return ""


from bodega.models import Movimiento, StockConfig, Vino
from proveedores.models import Proveedor, VinoProveedor
from pedidos.models import Pedido, LineaPedido


# ---------------------------------------------------------------------------
# Mapeos de secciones y proveedores
# ---------------------------------------------------------------------------

SECCIONES = {
    "Burbujas Nacional": Vino.Familia.ESPUMOSO_NAC,
    "Burbujas Internacional (Champagne)": Vino.Familia.CHAMPAGNE,
    "Burbujas Internacional (Otros)": Vino.Familia.ESPUMOSO_INT,
    "Blanco Nacional": Vino.Familia.BLANCO_NAC,
    "Blanco Internacional": Vino.Familia.BLANCO_INT,
    "Tintos Nacional": Vino.Familia.TINTO_NAC,
    "Tintos Internacional": Vino.Familia.TINTO_INT,
    "Grandes Vinos Tintos de España": Vino.Familia.GRANDES_TINTOS,
    "Rosados": Vino.Familia.ROSADO,
    "Dulces": Vino.Familia.DULCE,
    "Generosos": Vino.Familia.GENEROSO,
    "Fuera de Carta": Vino.Familia.FUERA_CARTA,
    "Vinos de Propiedad": Vino.Familia.PROPIEDAD,
}

HOJAS_PROVEEDOR = {
    "Coupa": "Coupa (Economato Meliá)",
    "Sanchez Polo": "Sánchez Polo",
    "Vila Viniteca": "Vila Viniteca",
    "LVMH": "LVMH (Moët Hennessy)",
    "Terroir Champenoise": "Terroir Champenoise",
    "Vega Tolosa": "Vega Tolosa",
    "Decantare": "Decantare",
    "Don Pepe": "Don Pepe",
    "Hibeta": "Hibeta",
    "Le Tribute (Pebar)": "Le Tribute (Pebar)",
    "Lopez Pardo": "López Pardo",
    "Narbona Solis": "Narbona Solís",
    "Vadevinos Exclusivas": "Vadevinos Exclusivas",
}


# ---------------------------------------------------------------------------
# Helpers de conversión
# ---------------------------------------------------------------------------

def safe_decimal(val, default="0"):
    if val == "" or val is None:
        return Decimal(default)
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def safe_int(val):
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_stock(val):
    if val == "" or val is None:
        return None
    try:
        return Decimal(str(val)).quantize(Decimal("0.1"))
    except (InvalidOperation, ValueError):
        return None


def safe_str(val):
    """Convierte valor de celda a str, devuelve '' para None/vacío."""
    if val is None or val == "":
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("none", "nan", "null") else s


def limpiar_nombre(nombre):
    """Limpia marcadores del nombre del vino: (?), (muestra), AGOTADO, etc."""
    nombre = nombre.strip()
    nombre = re.sub(r'^\s*\(\?\)\s*', '', nombre)          # (?) al inicio
    nombre = re.sub(r'\s*\(muestra\)\s*', '', nombre, flags=re.IGNORECASE)
    nombre = re.sub(r'\s*AGOTADO\s*', '', nombre, flags=re.IGNORECASE)
    return nombre.strip()


# ---------------------------------------------------------------------------
# Detección dinámica de columnas por cabecera
# ---------------------------------------------------------------------------

def leer_cabeceras(sh):
    """Lee la fila 0 y devuelve un dict {nombre_normalizado: índice_columna}."""
    hmap = {}
    for c in range(sh.ncols):
        h = str(sh.cell_value(0, c)).strip().lower()
        h = re.sub(r'[áàä]', 'a', h)
        h = re.sub(r'[éèë]', 'e', h)
        h = re.sub(r'[íìï]', 'i', h)
        h = re.sub(r'[óòö]', 'o', h)
        h = re.sub(r'[úùü]', 'u', h)
        h = re.sub(r'[ñ]', 'n', h)
        if h:
            hmap[h] = c
    return hmap


def col(hmap, *claves, default):
    """Busca el índice de columna probando varias claves posibles."""
    for clave in claves:
        clave_norm = clave.lower()
        if clave_norm in hmap:
            return hmap[clave_norm]
        # Búsqueda parcial solo para claves ≥4 chars para evitar falsos positivos
        # con abreviaturas de una letra (ej. 'c', 'n') que coinciden con cualquier cabecera
        if len(clave_norm) >= 4:
            for k, v in hmap.items():
                if len(k) >= 3 and (clave_norm in k or k in clave_norm):
                    return v
    return default


# ---------------------------------------------------------------------------
# Comando principal
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa datos del Libro de Bodega Excel"

    def add_arguments(self, parser):
        parser.add_argument("archivo", help="Ruta al archivo .xls/.xlsx del Libro de Bodega")

    def handle(self, *args, **options):
        path = options["archivo"]
        wb = abrir_libro(path)

        self.stdout.write("=" * 60)
        self.stdout.write("IMPORTADOR DE LIBRO DE BODEGA")
        self.stdout.write("=" * 60)

        proveedores = self._importar_proveedores(wb)
        vinos_map = self._importar_vinos(wb)
        self._vincular_proveedores(wb, vinos_map, proveedores)
        self._marcar_copas(wb, vinos_map)
        self._importar_pedidos(wb, vinos_map, proveedores)

        self.stdout.write("")
        self.stdout.write("=" * 60)
        self.stdout.write(self.style.SUCCESS(f"✓ {Vino.objects.count()} vinos importados"))
        self.stdout.write(self.style.SUCCESS(f"✓ {Proveedor.objects.count()} proveedores"))
        self.stdout.write(self.style.SUCCESS(f"✓ {Movimiento.objects.count()} movimientos de stock"))
        self.stdout.write(self.style.SUCCESS(f"✓ {Vino.objects.filter(es_copa=True).count()} vinos por copas"))
        bajo = sum(1 for v in Vino.objects.all() if v.bajo_minimo)
        self.stdout.write(self.style.SUCCESS(f"✓ {bajo} vinos bajo mínimo"))
        pedidos = Pedido.objects.count()
        if pedidos:
            self.stdout.write(self.style.SUCCESS(f"✓ {pedidos} pedidos pendientes importados"))

    # -----------------------------------------------------------------------

    def _importar_proveedores(self, wb):
        self.stdout.write("\n--- Importando proveedores ---")
        proveedores = {}

        for hoja_nombre, nombre_normalizado in HOJAS_PROVEEDOR.items():
            if hoja_nombre not in wb.sheet_names():
                continue
            sh = wb.sheet_by_name(hoja_nombre)
            if sh.nrows < 2:
                continue

            hmap = leer_cabeceras(sh)
            c_email    = col(hmap, 'email', 'correo', 'e-mail', default=10)
            c_contacto = col(hmap, 'contacto', 'nombre contacto', 'responsable', default=9)
            c_telefono = col(hmap, 'telefono', 'tlf', 'tel', default=11)

            email = contacto = telefono = ""
            for r in range(1, min(sh.nrows, 5)):
                e = safe_str(sh.cell_value(r, c_email)) if sh.ncols > c_email else ""
                if e and "@" in e:
                    email = e
                    contacto = safe_str(sh.cell_value(r, c_contacto)) if sh.ncols > c_contacto else ""
                    tel = sh.cell_value(r, c_telefono) if sh.ncols > c_telefono else ""
                    if tel:
                        telefono = str(int(tel)) if isinstance(tel, float) else safe_str(tel)
                    break

            prov = Proveedor.objects.create(
                nombre=nombre_normalizado,
                email=email,
                telefono=telefono,
                contacto=contacto,
            )
            proveedores[hoja_nombre] = prov
            self.stdout.write(f"  {nombre_normalizado} ({email})")

        return proveedores

    def _importar_vinos(self, wb):
        self.stdout.write("\n--- Importando vinos ---")
        sh = wb.sheet_by_name("Cañitas Maite (M)")

        hmap = leer_cabeceras(sh)
        self.stdout.write(f"  Cabeceras detectadas: {list(hmap.keys())[:10]}…")

        # Detectar columnas por nombre, con fallback a posición histórica
        c_nombre      = col(hmap, 'nombre', 'vino', 'articulo', 'producto', default=1)
        c_bodega      = col(hmap, 'bodega', 'productor', 'elaborador', default=2)
        c_azucar      = col(hmap, 'tipo', 'azucar', 'estilo', 'dulzor', default=3)
        c_do          = col(hmap, 'd.o.', 'denominacion', 'do', 'origen', default=4)
        c_variedades  = col(hmap, 'variedades', 'uva', 'uvas', 'cepa', default=5)
        c_precio      = col(hmap, 'precio coste', 'coste', 'p.coste', 'precio', default=6)
        c_stock       = col(hmap, 'stock', 'existencias', 'unidades', default=7)
        c_carta       = col(hmap, 'precio carta', 'pvp', 'p.v.p', 'venta', 'carta', default=8)
        c_distrib     = col(hmap, 'distribuidor', 'proveedor', 'distribucion', default=9)
        c_canitas     = col(hmap, 'canitas', 'c', default=10)
        c_ene         = col(hmap, 'ene', default=11)
        c_pool        = col(hmap, 'pool', default=12)

        vinos_map = {}
        familia_actual = Vino.Familia.ESPUMOSO_NAC
        count = 0

        for r in range(1, sh.nrows):
            nombre = safe_str(sh.cell_value(r, c_nombre))
            if not nombre or nombre == "TOTAL (sin IVA / con IVA)":
                continue

            # Detectar cabeceras de sección
            es_seccion = False
            for seccion_nombre, seccion_familia in SECCIONES.items():
                if nombre == seccion_nombre or nombre.startswith(seccion_nombre):
                    familia_actual = seccion_familia
                    es_seccion = True
                    break
            if es_seccion:
                continue

            nombre_limpio = limpiar_nombre(nombre)
            if not nombre_limpio or nombre_limpio in vinos_map:
                continue
            # Ignorar filas con "nombre" puramente numérico (referencias, totales, años…)
            if re.match(r'^[\d\s.,/-]+$', nombre_limpio):
                continue

            bodega       = safe_str(sh.cell_value(r, c_bodega))
            azucar       = safe_str(sh.cell_value(r, c_azucar))
            do           = safe_str(sh.cell_value(r, c_do))
            variedades   = safe_str(sh.cell_value(r, c_variedades))
            precio_coste = safe_decimal(sh.cell_value(r, c_precio))
            precio_carta = safe_decimal(sh.cell_value(r, c_carta))
            stock        = safe_stock(sh.cell_value(r, c_stock))
            distribuidor = safe_str(sh.cell_value(r, c_distrib))

            canitas_val  = safe_str(sh.cell_value(r, c_canitas)).upper()
            ene_val      = safe_str(sh.cell_value(r, c_ene)).upper()
            pool_val     = safe_str(sh.cell_value(r, c_pool)).upper()

            en_canitas = canitas_val in ("C", "B", "C-B")
            en_ene     = bool(ene_val) and ene_val not in ("", "0")
            en_pool    = pool_val in ("C", "OPCIONAL") or bool(pool_val and pool_val not in ("", "0"))
            via_coupa  = distribuidor.lower() == "coupa"

            vino = Vino.objects.create(
                nombre=nombre_limpio,
                bodega_nombre=bodega,
                familia=familia_actual,
                azucar=azucar,
                denominacion_origen=do,
                variedades=variedades,
                precio_coste=precio_coste,
                precio_carta=precio_carta,
                via_coupa=via_coupa,
                en_canitas=en_canitas,
                en_ene=en_ene,
                en_pool=en_pool,
            )

            vinos_map[nombre_limpio] = vino

            if stock is not None and stock > 0:
                Movimiento.objects.create(
                    vino=vino,
                    tipo=Movimiento.Tipo.ENTRADA,
                    cantidad=stock,
                    notas="Stock importado del Libro de Bodega",
                )

            StockConfig.objects.create(
                vino=vino,
                stock_minimo=6 if precio_coste < 50 else 2,
                stock_optimo=12 if precio_coste < 50 else 4,
            )

            vino._distribuidor_nombre = distribuidor
            count += 1

        self.stdout.write(f"  {count} vinos importados")
        return vinos_map

    def _vincular_proveedores(self, wb, vinos_map, proveedores):
        self.stdout.write("\n--- Vinculando vinos con proveedores ---")
        count = 0

        for hoja_nombre, prov in proveedores.items():
            if hoja_nombre not in wb.sheet_names():
                continue
            sh = wb.sheet_by_name(hoja_nombre)
            hmap = leer_cabeceras(sh)
            c_nombre = col(hmap, 'nombre', 'vino', 'articulo', 'producto', default=1)
            c_precio = col(hmap, 'precio', 'precio coste', 'coste', 'p.coste', default=4)

            for r in range(1, sh.nrows):
                nombre = safe_str(sh.cell_value(r, c_nombre))
                if not nombre or nombre == "TOTAL (sin IVA / con IVA)":
                    continue

                nombre_limpio = limpiar_nombre(nombre)
                vino = vinos_map.get(nombre_limpio) or _buscar_vino(vinos_map, nombre_limpio)
                if not vino:
                    continue

                precio = safe_decimal(sh.cell_value(r, c_precio) if sh.ncols > c_precio else 0)
                if not VinoProveedor.objects.filter(vino=vino, proveedor=prov).exists():
                    VinoProveedor.objects.create(vino=vino, proveedor=prov, precio=precio, es_principal=True)
                    count += 1

        self.stdout.write(f"  {count} vínculos vino-proveedor creados")

    def _marcar_copas(self, wb, vinos_map):
        self.stdout.write("\n--- Marcando vinos por copas ---")
        if "Vinos a Copas" not in wb.sheet_names():
            self.stdout.write("  Hoja 'Vinos a Copas' no encontrada")
            return

        sh = wb.sheet_by_name("Vinos a Copas")
        hmap = leer_cabeceras(sh)
        c_nombre     = col(hmap, 'nombre', 'vino', 'articulo', default=1)
        c_precio_copa = col(hmap, 'precio copa', 'precio/copa', 'copa', 'pvp copa', default=8)
        count = 0

        for r in range(1, sh.nrows):
            nombre = str(sh.cell_value(r, c_nombre)).strip()
            if not nombre or nombre == "TOTAL (sin IVA / con IVA)":
                continue

            nombre_limpio = limpiar_nombre(nombre)
            vino = vinos_map.get(nombre_limpio) or _buscar_vino(vinos_map, nombre_limpio)
            if not vino:
                continue

            precio_copa = safe_decimal(sh.cell_value(r, c_precio_copa) if sh.ncols > c_precio_copa else 0)
            vino.es_copa = True
            vino.precio_copa = precio_copa
            vino.save()
            count += 1

        self.stdout.write(f"  {count} vinos marcados como copa")

    def _importar_pedidos(self, wb, vinos_map, proveedores):
        self.stdout.write("\n--- Importando pedidos pendientes ---")

        for hoja_nombre, prov in proveedores.items():
            if hoja_nombre not in wb.sheet_names():
                continue
            sh = wb.sheet_by_name(hoja_nombre)
            hmap = leer_cabeceras(sh)
            c_nombre  = col(hmap, 'nombre', 'vino', 'articulo', default=1)
            c_pedido  = col(hmap, 'pedido', 'cantidad pedido', 'a pedir', default=3)
            c_precio  = col(hmap, 'precio', 'precio coste', 'coste', default=4)

            lineas_pedido = []
            for r in range(1, sh.nrows):
                nombre = str(sh.cell_value(r, c_nombre)).strip()
                if not nombre or nombre == "TOTAL (sin IVA / con IVA)":
                    continue

                pedido_val = sh.cell_value(r, c_pedido) if sh.ncols > c_pedido else ""
                if not pedido_val or pedido_val == "":
                    continue
                try:
                    cantidad = int(float(pedido_val))
                except (ValueError, TypeError):
                    continue
                if cantidad <= 0:
                    continue

                nombre_limpio = limpiar_nombre(nombre)
                vino = vinos_map.get(nombre_limpio) or _buscar_vino(vinos_map, nombre_limpio)
                if not vino:
                    continue

                precio = safe_decimal(sh.cell_value(r, c_precio) if sh.ncols > c_precio else 0)
                lineas_pedido.append({"vino": vino, "cantidad": cantidad, "precio": precio})

            if lineas_pedido:
                pedido = Pedido.objects.create(
                    proveedor=prov,
                    estado=Pedido.Estado.BORRADOR,
                    generado_por_ia=False,
                    notas=f"Pedido importado del Libro de Bodega (hoja {hoja_nombre})",
                )
                for lp in lineas_pedido:
                    LineaPedido.objects.create(
                        pedido=pedido,
                        vino=lp["vino"],
                        cantidad_sugerida=lp["cantidad"],
                        precio_unitario=lp["precio"],
                    )
                self.stdout.write(f"  Pedido {prov.nombre}: {len(lineas_pedido)} líneas")

def _buscar_vino(vinos_map, nombre):
    """Búsqueda tolerante: coincidencia parcial bidireccional."""
    nl = nombre.lower()
    for key, v in vinos_map.items():
        kl = key.lower()
        if nl in kl or kl in nl:
            return v
    return None
