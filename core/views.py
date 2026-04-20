from decimal import Decimal

import io
import tempfile
import os

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.core.management import call_command
from django.shortcuts import render, redirect
from django.http import HttpResponseForbidden
from django.db.models import Sum, DecimalField, Value
from django.db.models.functions import Coalesce

from django.http import JsonResponse
from django.views.decorators.http import require_POST

from bodega.models import Vino, StockConfig, Movimiento
from pedidos.models import Pedido, LineaPedido
from proveedores.models import Proveedor, VinoProveedor
from .forms import RegistroForm, PerfilForm, AvatarForm
from .models import Anotacion, PerfilUsuario, Configuracion


@login_required
def dashboard(request):
    # Una sola query: stock calculado + config de mínimos
    vinos = list(
        Vino.objects
        .filter(activo=True)
        .select_related("stock_config")
        .annotate(
            stock_anotado=Coalesce(
                Sum("movimientos__cantidad"),
                Value(Decimal("0")),
                output_field=DecimalField(),
            )
        )
    )

    # Resumen por familia (en Python, sin más queries)
    familias = []
    for codigo, nombre in Vino.Familia.choices:
        vinos_familia = [v for v in vinos if v.familia == codigo]
        if not vinos_familia:
            continue
        stock_total = sum(v.stock_anotado for v in vinos_familia)
        bajo_minimo_count = sum(
            1 for v in vinos_familia
            if hasattr(v, "stock_config") and v.stock_anotado < v.stock_config.stock_minimo
        )
        familias.append({
            "codigo": codigo,
            "nombre": nombre,
            "count": len(vinos_familia),
            "stock_total": stock_total,
            "bajo_minimo": bajo_minimo_count,
        })

    # Alertas: vinos bajo mínimo (sin queries adicionales)
    alertas = []
    for v in vinos:
        try:
            if v.stock_anotado < v.stock_config.stock_minimo:
                alertas.append({
                    "vino": v,
                    "stock_actual": v.stock_anotado,
                    "stock_minimo": v.stock_config.stock_minimo,
                })
        except StockConfig.DoesNotExist:
            pass

    pedidos_pendientes = Pedido.objects.filter(
        estado__in=[Pedido.Estado.BORRADOR, Pedido.Estado.PENDIENTE]
    ).count()

    valor_inventario = sum(v.stock_anotado * v.precio_coste for v in vinos)

    chart_labels = [f["nombre"] for f in familias]
    chart_data = [float(f["stock_total"]) for f in familias]

    context = {
        "familias": familias,
        "alertas": alertas,
        "pedidos_pendientes": pedidos_pendientes,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "total_vinos": len(vinos),
        "valor_inventario": valor_inventario,
    }
    return render(request, "dashboard.html", context)


@login_required
def ayuda(request):
    return render(request, "ayuda.html")


@login_required
def anotaciones(request):
    if request.method == "POST":
        texto = request.POST.get("texto", "").strip()
        prioridad = request.POST.get("prioridad", Anotacion.Prioridad.NORMAL)
        if texto:
            Anotacion.objects.create(texto=texto, prioridad=prioridad)
        return redirect("core:anotaciones")

    lista = Anotacion.objects.filter(resuelta=False)
    resueltas = Anotacion.objects.filter(resuelta=True)[:10]
    return render(request, "anotaciones.html", {"lista": lista, "resueltas": resueltas})


@login_required
@require_POST
def anotacion_resolver(request, pk):
    Anotacion.objects.filter(pk=pk).update(resuelta=True)
    return JsonResponse({"ok": True})


@login_required
@require_POST
def anotacion_eliminar(request, pk):
    Anotacion.objects.filter(pk=pk).delete()
    return JsonResponse({"ok": True})


def _importar_plantilla_generica(archivo):
    import openpyxl
    from decimal import Decimal, InvalidOperation
    from bodega.models import Vino, StockConfig, Movimiento
    from proveedores.models import Proveedor, VinoProveedor

    FAMILIA_MAP = {label.lower(): key for key, label in Vino.Familia.choices}

    def to_dec(v, default="0"):
        try:
            return Decimal(str(v)).quantize(Decimal("0.01"))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    def to_int(v, default=0):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default

    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("El archivo no contiene datos.")

    headers = [str(h).strip().lower().rstrip(" *") if h else "" for h in rows[0]]

    def col(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        return None

    c_nombre   = col("nombre")
    c_bodega   = col("bodega")
    c_tipo     = col("tipo")
    c_do       = col("d.o")
    c_var      = col("variedad")
    c_pcoste   = col("precio coste")
    c_pcarta   = col("precio carta")
    c_stock    = col("stock inicial")
    c_smin     = col("stock mín")
    c_sopt     = col("stock óptimo")
    c_prov     = col("proveedor")
    c_email    = col("email")
    c_copa     = col("copas")
    c_pcopa    = col("precio copa")

    if c_nombre is None or c_tipo is None:
        raise ValueError("Faltan columnas obligatorias: «Nombre» y «Tipo».")

    vinos_creados = provs_creados = 0
    proveedores_cache = {}

    for row in rows[1:]:
        nombre = str(row[c_nombre]).strip() if row[c_nombre] else ""
        if not nombre or nombre.lower() in ("none", ""):
            continue

        tipo_raw = str(row[c_tipo]).strip().lower() if c_tipo and row[c_tipo] else ""
        familia = FAMILIA_MAP.get(tipo_raw)
        if not familia:
            familia = Vino.Familia.TINTO_NAC

        def get(c):
            return str(row[c]).strip() if c is not None and row[c] not in (None, "") else ""

        vino = Vino.objects.create(
            nombre=nombre,
            bodega_nombre=get(c_bodega),
            familia=familia,
            denominacion_origen=get(c_do),
            variedades=get(c_var),
            precio_coste=to_dec(get(c_pcoste)),
            precio_carta=to_dec(get(c_pcarta)),
            es_copa=get(c_copa).lower() in ("sí", "si", "s", "yes", "1") if c_copa else False,
            precio_copa=to_dec(get(c_pcopa)) if c_pcopa else Decimal("0"),
        )

        stock_ini = to_int(get(c_stock))
        if stock_ini > 0:
            Movimiento.objects.create(
                vino=vino, tipo=Movimiento.Tipo.ENTRADA,
                cantidad=stock_ini, notas="Stock importado de plantilla"
            )

        StockConfig.objects.create(
            vino=vino,
            stock_minimo=to_int(get(c_smin)) or 6,
            stock_optimo=to_int(get(c_sopt)) or 12,
        )

        prov_nombre = get(c_prov)
        prov_email = get(c_email)
        if prov_nombre:
            if prov_nombre not in proveedores_cache:
                prov, created = Proveedor.objects.get_or_create(
                    nombre=prov_nombre,
                    defaults={"email": prov_email}
                )
                if created:
                    provs_creados += 1
                proveedores_cache[prov_nombre] = prov
            VinoProveedor.objects.get_or_create(
                vino=vino, proveedor=proveedores_cache[prov_nombre],
                defaults={"precio": to_dec(get(c_pcoste)), "es_principal": True}
            )

        vinos_creados += 1

    return f"Importación completada: {vinos_creados} vinos y {provs_creados} proveedores nuevos."


@login_required
def herramientas(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Solo superusuarios.")

    stats = {
        "vinos": Vino.objects.count(),
        "movimientos": Movimiento.objects.count(),
        "proveedores": Proveedor.objects.count(),
        "pedidos": Pedido.objects.count(),
        "anotaciones": Anotacion.objects.count(),
        "vinos_sin_imagen": Vino.objects.filter(activo=True, imagen="").count(),
    }

    if request.method == "POST":
        accion = request.POST.get("accion")

        if accion == "borrar_imagenes":
            vinos_con_imagen = Vino.objects.exclude(imagen="")
            count = 0
            for v in vinos_con_imagen:
                if v.imagen:
                    v.imagen.delete(save=False)
                    v.imagen = ""
                    v.save(update_fields=["imagen"])
                    count += 1
            messages.success(request, f"{count} imágenes eliminadas. Puedes volver a escanear.")
            return redirect("core:herramientas")

        elif accion == "limpiar":
            LineaPedido.objects.all().delete()
            Pedido.objects.all().delete()
            VinoProveedor.objects.all().delete()
            Movimiento.objects.all().delete()
            StockConfig.objects.all().delete()
            Vino.objects.all().delete()
            Proveedor.objects.all().delete()
            Anotacion.objects.all().delete()
            messages.success(request, "Base de datos limpiada. Puedes importar un Excel nuevo.")
            return redirect("core:herramientas")

        elif accion == "importar":
            archivo = request.FILES.get("excel")
            if not archivo:
                messages.error(request, "Selecciona un archivo.")
                return redirect("core:herramientas")

            nombre = archivo.name.lower()
            if nombre.endswith(".xlsx"):
                sufijo = ".xlsx"
            elif nombre.endswith(".xls"):
                sufijo = ".xls"
            else:
                messages.error(request, "Formato no soportado. Usa .xls o .xlsx (Excel / LibreOffice / OpenOffice).")
                return redirect("core:herramientas")

            with tempfile.NamedTemporaryFile(suffix=sufijo, delete=False) as tmp:
                for chunk in archivo.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            try:
                out = io.StringIO()
                call_command("importar_excel", tmp_path, stdout=out, stderr=out)
                output = out.getvalue()
                vinos_nuevos = Vino.objects.count()
                messages.success(
                    request,
                    f"Importación completada: {vinos_nuevos} vinos, "
                    f"{Proveedor.objects.count()} proveedores, "
                    f"{Movimiento.objects.count()} movimientos."
                )
            except Exception as e:
                messages.error(request, f"Error al importar: {e}")
            finally:
                os.unlink(tmp_path)

            return redirect("core:herramientas")

        elif accion == "importar_plantilla":
            archivo = request.FILES.get("plantilla")
            if not archivo:
                messages.error(request, "Selecciona un archivo.")
                return redirect("core:herramientas")
            try:
                resultado = _importar_plantilla_generica(archivo)
                messages.success(request, resultado)
            except Exception as e:
                messages.error(request, f"Error al importar plantilla: {e}")
            return redirect("core:herramientas")

        elif accion == "logo":
            config = Configuracion.get()
            if request.FILES.get("logo"):
                config.logo = request.FILES["logo"]
                config.save()
                messages.success(request, "Logo actualizado.")
            return redirect("core:herramientas")

        elif accion == "login_imagen":
            config = Configuracion.get()
            if request.FILES.get("login_imagen"):
                config.login_imagen = request.FILES["login_imagen"]
                config.save()
                messages.success(request, "Imagen de login actualizada.")
            return redirect("core:herramientas")
        elif accion == "borrar_logo":
            config = Configuracion.get()
            if config.logo:
                config.logo.delete(save=False)
                config.logo = ""
                config.save()
                messages.success(request, "Logo eliminado. Se mostrará la imagen por defecto.")
            return redirect("core:herramientas")
        elif accion == "borrar_login_imagen":
            config = Configuracion.get()
            if config.login_imagen:
                config.login_imagen.delete(save=False)
                config.login_imagen = ""
                config.save()
                messages.success(request, "Imagen de login eliminada. Se mostrará la imagen por defecto.")
            return redirect("core:herramientas")

    return render(request, "herramientas.html", {"stats": stats})


@login_required
def descargar_plantilla(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from bodega.models import Vino

    wb = openpyxl.Workbook()

    # ── Hoja principal: Vinos ──
    ws = wb.active
    ws.title = "Vinos"

    headers = [
        "Nombre *", "Bodega", "Tipo *", "D.O.",
        "Variedades", "Precio Coste", "Precio Carta",
        "Stock Inicial", "Stock Mínimo", "Stock Óptimo",
        "Proveedor", "Email Proveedor",
        "Se vende por copas (Sí/No)", "Precio Copa",
    ]

    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Filas de ejemplo
    ejemplos = [
        ["Ribera del Duero Crianza", "Bodega Pesquera", "Tinto Nacional",
         "Ribera del Duero", "Tempranillo", 12.50, 38.00, 18, 6, 12,
         "Distribuciones García", "pedidos@garcia.com", "No", ""],
        ["Albariño Mar de Frades", "Mar de Frades", "Blanco Nacional",
         "Rías Baixas", "Albariño", 9.80, 32.00, 12, 6, 12,
         "Vinos del Norte", "info@vinosnorte.com", "Sí", 8.50],
        ["Moët & Chandon Brut", "Moët & Chandon", "Champagne",
         "Champagne", "Pinot Noir, Chardonnay, Pinot Meunier", 38.00, 95.00, 6, 2, 6,
         "LVMH", "lvmh@example.com", "No", ""],
    ]
    ex_fill = PatternFill("solid", fgColor="F1F5F9")
    for r, row in enumerate(ejemplos, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = ex_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Anchos de columna
    col_widths = [30, 22, 24, 22, 28, 14, 14, 14, 14, 14, 24, 26, 26, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.row_dimensions[1].height = 30

    # ── Hoja de referencia: Tipos válidos ──
    ws2 = wb.create_sheet("Tipos válidos")
    ws2.cell(1, 1, "Valores aceptados en la columna «Tipo»")
    ws2.cell(1, 1).font = Font(bold=True, size=11)
    ws2.column_dimensions["A"].width = 40
    for i, (_, label) in enumerate(Vino.Familia.choices, 2):
        ws2.cell(i, 1, label)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_bodega_someliar.xlsx"'
    wb.save(response)
    return response


def registro(request):
    if request.user.is_authenticated:
        return redirect("core:dashboard")
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            from django.contrib.auth.models import User
            user = form.save(commit=False)
            if not User.objects.exists():
                user.is_staff = True
                user.is_superuser = True
            user.save()
            login(request, user)
            messages.success(request, f"¡Bienvenido, {user.first_name or user.username}! Cuenta creada correctamente.")
            return redirect("core:dashboard")
    else:
        form = RegistroForm()
    return render(request, "registration/register.html", {"form": form})


@login_required
def perfil(request):
    perfil_obj, _ = PerfilUsuario.objects.get_or_create(user=request.user)

    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion == "avatar":
            avatar_form = AvatarForm(request.POST, request.FILES, instance=perfil_obj)
            if avatar_form.is_valid():
                avatar_form.save()
                messages.success(request, "Foto de perfil actualizada.")
            return redirect("core:perfil")
        elif accion == "borrar_avatar":
            if perfil_obj.avatar:
                perfil_obj.avatar.delete(save=False)
                perfil_obj.avatar = ""
                perfil_obj.save()
                messages.success(request, "Foto eliminada.")
            return redirect("core:perfil")
        else:
            form = PerfilForm(request.POST, instance=request.user)
            if form.is_valid():
                form.save()
                perfil_obj.foto_en_sidebar = "foto_en_sidebar" in request.POST
                perfil_obj.save()
                messages.success(request, "Perfil actualizado correctamente.")
                return redirect("core:perfil")
            avatar_form = AvatarForm(instance=perfil_obj)
            return render(request, "registration/profile.html", {"form": form, "avatar_form": avatar_form})

    form = PerfilForm(instance=request.user)
    avatar_form = AvatarForm(instance=perfil_obj)
    return render(request, "registration/profile.html", {"form": form, "avatar_form": avatar_form})




